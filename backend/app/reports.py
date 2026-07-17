import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_csv_report(validation_results: list) -> str:
    """
    Generates a CSV formatted string of matching validation results.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Prescribed Medicine", "Prescribed Strength", "Billed Item", "Billed Strength", 
        "Billed Qty", "Match Score", "Decision Status", "Match Reason"
    ])
    
    for r in validation_results:
        p_name = r.get("prescribed_name") or "N/A"
        p_strength = r.get("prescribed_strength") or "N/A"
        b_name = r.get("billed_name") or "N/A"
        b_strength = r.get("billed_strength") or "N/A"
        b_qty = r.get("billed_qty") or "N/A"
        score = f"{r.get('match_score', 0):.1f}%"
        status = r.get("match_status") or "Pending"
        reason = r.get("match_reason") or ""
        
        writer.writerow([p_name, p_strength, b_name, b_strength, b_qty, score, status, reason])
        
    return output.getvalue()

def generate_excel_report(claim_data: dict, validation_results: list) -> io.BytesIO:
    """
    Generates a styled Excel sheet using openpyxl.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Claim Audit Report"
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    font_title = Font(name="Calibri", size=16, bold=True, color="FF6F00") # IOCL Orange
    font_section = Font(name="Calibri", size=12, bold=True)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    
    fill_header = PatternFill(start_color="0A192F", end_color="0A192F", fill_type="solid") # Dark Navy
    fill_approved = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # Soft Green
    fill_rejected = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red
    fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Soft Yellow
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # 1. Header block
    ws.merge_cells("A1:H1")
    ws["A1"] = "INDIAN OIL CORPORATION LIMITED - MEDICAL CLAIM AUDIT REPORT"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # 2. Metadata block
    ws["A3"] = "Claim ID:"
    ws["A3"].font = font_bold
    ws["B3"] = claim_data.get("claim_id")
    ws["B3"].font = font_regular
    
    ws["A4"] = "Employee Name:"
    ws["A4"].font = font_bold
    ws["B4"] = claim_data.get("employee_name")
    ws["B4"].font = font_regular
    
    ws["A5"] = "Department:"
    ws["A5"].font = font_bold
    ws["B5"] = claim_data.get("department")
    ws["B5"].font = font_regular
    
    ws["E3"] = "Date Submitted:"
    ws["E3"].font = font_bold
    ws["F3"] = claim_data.get("date_submitted")
    ws["F3"].font = font_regular
    
    ws["E4"] = "Audit Status:"
    ws["E4"].font = font_bold
    ws["F4"] = claim_data.get("status")
    ws["F4"].font = font_regular
    
    # Summary amounts
    ws["A7"] = "CLAIM SUMMARY"
    ws["A7"].font = font_section
    
    ws["A8"] = "Total Claimed Amount:"
    ws["A8"].font = font_bold
    ws["B8"] = f"INR {claim_data.get('claimed_amount', 0.0):,.2f}"
    ws["B8"].font = font_regular
    
    ws["D8"] = "Approved Amount:"
    ws["D8"].font = font_bold
    ws["E8"] = f"INR {claim_data.get('approved_amount', 0.0):,.2f}"
    ws["E8"].font = font_regular
    
    ws["G8"] = "Rejected Amount:"
    ws["G8"].font = font_bold
    ws["H8"] = f"INR {claim_data.get('rejected_amount', 0.0):,.2f}"
    ws["H8"].font = font_regular
    
    # Table headers
    headers = [
        "Prescribed Medicine", "Prescribed Strength", "Billed Item", "Billed Strength", 
        "Billed Qty", "Match Score", "Decision Status", "Reason"
    ]
    
    start_row = 10
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border
        
    ws.row_dimensions[start_row].height = 28
    
    # Seed table rows
    current_row = start_row + 1
    for r in validation_results:
        ws.cell(row=current_row, column=1, value=r.get("prescribed_name") or "N/A").font = font_regular
        ws.cell(row=current_row, column=2, value=r.get("prescribed_strength") or "N/A").font = font_regular
        ws.cell(row=current_row, column=3, value=r.get("billed_name") or "N/A").font = font_regular
        ws.cell(row=current_row, column=4, value=r.get("billed_strength") or "N/A").font = font_regular
        ws.cell(row=current_row, column=5, value=r.get("billed_qty") or 0).font = font_regular
        
        score_cell = ws.cell(row=current_row, column=6, value=f"{r.get('match_score', 0):.1f}%")
        score_cell.font = font_regular
        score_cell.alignment = Alignment(horizontal="center")
        
        status_cell = ws.cell(row=current_row, column=7, value=r.get("match_status"))
        status_cell.font = font_bold
        status_cell.alignment = Alignment(horizontal="center")
        
        # Apply conditional fills based on status
        if r.get("match_status") == "Approved":
            status_cell.fill = fill_approved
        elif r.get("match_status") == "Rejected":
            status_cell.fill = fill_rejected
        else:
            status_cell.fill = fill_pending
            
        ws.cell(row=current_row, column=8, value=r.get("match_reason") or "").font = font_regular
        
        # Apply borders to row
        for col_idx in range(1, 9):
            ws.cell(row=current_row, column=col_idx).border = box_border
            
        current_row += 1
        
    # Adjust column widths dynamically
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            # Cells from merged ranges (e.g. the title banner) have no column_letter — skip them
            if not hasattr(cell, "column_letter"):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

def generate_pdf_report(claim_data: dict, validation_results: list) -> io.BytesIO:
    """
    Generates a high-quality PDF report using ReportLab.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    style_title = ParagraphStyle(
        name='IOCLTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#0A192F'), # Dark Blue
        spaceAfter=15,
        alignment=1 # Centered
    )
    
    style_section = ParagraphStyle(
        name='IOCLSection',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#FF6F00'), # Saffron
        spaceBefore=12,
        spaceAfter=6
    )
    
    style_meta_label = ParagraphStyle(
        name='IOCLMetaLabel',
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12
    )
    
    style_meta_val = ParagraphStyle(
        name='IOCLMetaVal',
        fontName='Helvetica',
        fontSize=9,
        leading=12
    )
    
    style_th = ParagraphStyle(
        name='IOCLTableHeader',
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1
    )
    
    style_td = ParagraphStyle(
        name='IOCLTableCell',
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    style_td_status = ParagraphStyle(
        name='IOCLTableStatus',
        fontName='Helvetica-Bold',
        fontSize=8,
        alignment=1
    )

    # 1. Header Title
    story.append(Paragraph("INDIAN OIL CORPORATION LIMITED", style_title))
    story.append(Paragraph("Medical Reimbursement Audit Report", ParagraphStyle(
        name='Sub', fontName='Helvetica-Oblique', fontSize=12, alignment=1, textColor=colors.HexColor('#555555'), spaceAfter=20
    )))
    
    # 2. Metadata Grid
    meta_data = [
        [
            Paragraph("Claim ID:", style_meta_label), Paragraph(str(claim_data.get("claim_id")), style_meta_val),
            Paragraph("Date Submitted:", style_meta_label), Paragraph(str(claim_data.get("date_submitted")), style_meta_val)
        ],
        [
            Paragraph("Employee Name:", style_meta_label), Paragraph(str(claim_data.get("employee_name")), style_meta_val),
            Paragraph("Claim Status:", style_meta_label), Paragraph(str(claim_data.get("status")), style_meta_val)
        ],
        [
            Paragraph("Department:", style_meta_label), Paragraph(str(claim_data.get("department")), style_meta_val),
            Paragraph("Reviewer:", style_meta_label), Paragraph(str(claim_data.get("reviewer_name") or "System"), style_meta_val)
        ]
    ]
    
    meta_table = Table(meta_data, colWidths=[100, 150, 100, 150])
    meta_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#EAEAEA'))
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 15))
    
    # 3. Financial Summary
    story.append(Paragraph("FINANCIAL SUMMARY", style_section))
    financial_data = [
        [
            Paragraph("Total Claimed:", style_meta_label), Paragraph(f"INR {claim_data.get('claimed_amount', 0.0):,.2f}", style_meta_val),
            Paragraph("Approved Amount:", style_meta_label), Paragraph(f"INR {claim_data.get('approved_amount', 0.0):,.2f}", style_meta_val),
            Paragraph("Rejected Amount:", style_meta_label), Paragraph(f"INR {claim_data.get('rejected_amount', 0.0):,.2f}", style_meta_val)
        ]
    ]
    
    fin_table = Table(financial_data, colWidths=[90, 90, 100, 90, 100, 90])
    fin_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8F9FA')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#D9D9D9')),
        ('PADDING', (0, 0), (-1, -1), 8)
    ]))
    story.append(fin_table)
    story.append(Spacer(1, 15))
    
    # 4. Medicine Comparison Table
    story.append(Paragraph("MEDICINE VALIDATION DETAIL", style_section))
    
    # Setup headers
    table_content = [[
        Paragraph("Prescribed Item", style_th),
        Paragraph("Strength", style_th),
        Paragraph("Billed Item", style_th),
        Paragraph("Billed Str", style_th),
        Paragraph("Qty", style_th),
        Paragraph("Score", style_th),
        Paragraph("Decision", style_th),
        Paragraph("Reason / Rules Applied", style_th)
    ]]
    
    # Populate rows
    row_styles = [] # Used for background coloring of decisions
    for idx, r in enumerate(validation_results):
        p_name = r.get("prescribed_name") or "N/A"
        p_strength = r.get("prescribed_strength") or "N/A"
        b_name = r.get("billed_name") or "N/A"
        b_strength = r.get("billed_strength") or "N/A"
        b_qty = str(r.get("billed_qty") or 0)
        score = f"{r.get('match_score', 0):.1f}%"
        status = r.get("match_status")
        reason = r.get("match_reason") or ""
        
        status_color = "#E2F0D9" # Light green
        if status == "Rejected":
            status_color = "#FCE4D6" # Light red
        elif status == "Pending Review":
            status_color = "#FFF2CC" # Light yellow
            
        table_content.append([
            Paragraph(p_name, style_td),
            Paragraph(p_strength, style_td),
            Paragraph(b_name, style_td),
            Paragraph(b_strength, style_td),
            Paragraph(b_qty, style_td),
            Paragraph(score, style_td),
            Paragraph(status, style_td_status),
            Paragraph(reason, style_td)
        ])
        
        # Save color instruction: row, col, color hex
        row_styles.append((idx + 1, status_color))
        
    comp_table = Table(table_content, colWidths=[90, 50, 90, 50, 30, 40, 60, 120])
    
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0A192F')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]
    
    # Apply row status cell backgrounds
    for r_idx, c_hex in row_styles:
        t_style.append(('BACKGROUND', (6, r_idx), (6, r_idx), colors.HexColor(c_hex)))
        
    comp_table.setStyle(TableStyle(t_style))
    story.append(comp_table)
    
    story.append(Spacer(1, 30))
    
    # 5. Signatures and disclaimer
    sig_data = [
        [Paragraph("Prepared by: System AI Auditor", style_meta_label), Paragraph("Approved by: Medical Claims Reviewer / CMO", style_meta_label)],
        ["", ""], # Space for signature
        [Paragraph("Signature: _______________________", style_meta_label), Paragraph("Signature: _______________________", style_meta_label)],
        [Paragraph(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}", style_meta_val), Paragraph("Date: ___________________________", style_meta_val)]
    ]
    sig_table = Table(sig_data, colWidths=[270, 270])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(sig_table)
    
    # Build document
    doc.build(story)
    buffer.seek(0)
    return buffer
